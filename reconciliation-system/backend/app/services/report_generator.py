"""
Report Generator Service
Generates reports by filling Excel templates with reconciliation results
"""

import json
import logging
from pathlib import Path
from typing import Dict, Any, List, Optional
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

from app.core.config import get_storage_path
from app.core.database import DatabaseManager
from app.core.exceptions import ConfigurationError
from app.utils.file_utils import get_export_path
from app.utils.excel_utils import dataframe_to_csv

# Setup logger - use 'report' logger for dedicated log file
logger = logging.getLogger('report')


class ReportGenerator:
    """
    Generates reports from reconciliation results
    - Creates temp table from A1 data
    - Executes SQL queries on temp table
    - Fills Excel template with results
    - Cleans up temp table
    """
    
    def __init__(
        self,
        partner_code: str,
        service_code: str,
        period: str,
        batch_id: str,
        period_from,
        period_to,
        created_by: str = ""
    ):
        self.partner_code = partner_code
        self.service_code = service_code
        self.period = period
        self.batch_id = batch_id
        self.period_from = period_from
        self.period_to = period_to
        self.created_by = created_by
        self.temp_table_name = f"temp_a1_{batch_id.replace('-', '_').replace('.', '_')[:30]}"
        self.export_path = get_export_path(partner_code, period, batch_id)
        
        # Execution logs for tracking
        self.execution_logs = []
    
    def log_step(self, step: str, status: str, message: str, sql: str = None, result: Any = None):
        """Add a step log entry"""
        log_entry = {
            "step": step,
            "time": datetime.now().isoformat(),
            "status": status,
            "message": message
        }
        if sql:
            log_entry["sql"] = sql
        if result is not None:
            log_entry["result"] = str(result)[:500]  # Truncate long results
        
        self.execution_logs.append(log_entry)
        
        # Also log to file
        if status == "error":
            logger.error(f"[{step}] {message} | SQL: {sql}")
        else:
            logger.info(f"[{step}] {message}")
    
    def get_logs(self) -> List[Dict]:
        """Get all execution logs"""
        return self.execution_logs
    
    def generate_report(
        self,
        a1_df: pd.DataFrame,
        template_path: Optional[str],
        cell_mapping: Optional[Dict[str, Any]]
    ) -> Optional[str]:
        """
        Generate report from A1 data
        
        Args:
            a1_df: A1 result DataFrame
            template_path: Path to Excel template (relative to templates/)
            cell_mapping: Cell mapping configuration
        
        Returns:
            Path to generated report or None if no template
        """
        self.log_step("init", "start", f"Starting report generation for batch {self.batch_id}")
        
        if not template_path or not cell_mapping:
            self.log_step("init", "error", "Missing template_path or cell_mapping")
            return None
        
        try:
            # Create temp table
            self.log_step("create_temp_table", "start", f"Creating temp table: {self.temp_table_name}")
            self.log_step("create_temp_table", "info", f"A1 DataFrame has {len(a1_df)} rows, columns: {list(a1_df.columns)}")
            self._create_temp_table(a1_df)
            self.log_step("create_temp_table", "ok", f"Temp table created successfully")
            
            # Get template full path
            templates_dir = get_storage_path('templates')
            template_full_path = templates_dir / template_path
            
            if not template_full_path.exists():
                self.log_step("load_template", "error", f"Template not found: {template_path}")
                raise ConfigurationError(f"Template not found: {template_path}")
            
            self.log_step("load_template", "ok", f"Loaded template: {template_path}")
            
            # Load workbook
            wb = load_workbook(template_full_path)
            self.log_step("load_template", "info", f"Workbook sheets: {wb.sheetnames}")
            
            # Prepare static data
            static_data = {
                'partner_code': self.partner_code,
                'partner_name': self.partner_code,  # Could be looked up
                'service_code': self.service_code,
                'service_name': self.service_code,  # Could be looked up
                'period_from': str(self.period_from),
                'period_to': str(self.period_to),
                'created_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'created_by': self.created_by,
                'batch_id': self.batch_id
            }
            
            # Process each sheet
            sheets_config = cell_mapping.get('sheets', [])
            
            for sheet_config in sheets_config:
                self._process_sheet(wb, sheet_config, static_data)
            
            # Save output
            output_filename = f"Report_{self.partner_code}_{self.service_code}_{self.period}.xlsx"
            output_path = self.export_path / output_filename
            
            wb.save(output_path)
            
            return str(output_path)
        
        finally:
            # Cleanup temp table
            self._drop_temp_table()
    
    def _create_temp_table(self, a1_df: pd.DataFrame):
        """Create temporary table from A1 DataFrame"""
        DatabaseManager.create_temp_table(self.temp_table_name, a1_df)
    
    def _drop_temp_table(self):
        """Drop temporary table"""
        try:
            DatabaseManager.drop_temp_table(self.temp_table_name)
            self.log_step("cleanup", "ok", f"Dropped temp table: {self.temp_table_name}")
        except Exception as e:
            self.log_step("cleanup", "warning", f"Failed to drop temp table: {str(e)}")
    
    def _process_sheet(
        self,
        wb,
        sheet_config: Dict[str, Any],
        static_data: Dict[str, str]
    ):
        """Process a single sheet in the workbook"""
        sheet_name = sheet_config.get('sheet_name')
        
        # Create sheet if not exists
        if sheet_name not in wb.sheetnames:
            self.log_step("process_sheet", "info", f"Sheet '{sheet_name}' not found, creating new sheet")
            wb.create_sheet(title=sheet_name)
        
        self.log_step("process_sheet", "start", f"Processing sheet: {sheet_name}")
        ws = wb[sheet_name]
        
        # Fill static cells
        static_cells = sheet_config.get('static_cells', {})
        for cell_ref, value_template in static_cells.items():
            value = value_template
            for key, val in static_data.items():
                value = value.replace(f"{{{key}}}", str(val))
            ws[cell_ref] = value
            self.log_step("static_cell", "ok", f"Cell {cell_ref} = '{value}'")
        
        # Execute SQL and fill cells
        sql_cells = sheet_config.get('sql_cells', [])
        for sql_cell in sql_cells:
            cell_ref = sql_cell.get('cell')
            sql = sql_cell.get('sql', '')
            
            # Replace temp table name placeholder
            sql_executed = sql.replace('temp_a1', self.temp_table_name)
            
            try:
                results = DatabaseManager.execute_sql_on_temp(sql_executed)
                if results:
                    # Get first column of first row
                    first_row = results[0]
                    value = list(first_row.values())[0]
                    ws[cell_ref] = value
                    self.log_step("sql_cell", "ok", f"Cell {cell_ref} = {value}", sql=sql_executed, result=value)
                else:
                    ws[cell_ref] = 0
                    self.log_step("sql_cell", "warning", f"Cell {cell_ref} - No results", sql=sql_executed, result="empty")
            except Exception as e:
                error_msg = str(e)
                ws[cell_ref] = f"ERROR"
                self.log_step("sql_cell", "error", f"Cell {cell_ref} - {error_msg}", sql=sql_executed, result=error_msg)
        
        # Fill data table
        data_start_cell = sheet_config.get('data_start_cell')
        data_sql = sheet_config.get('data_sql')
        columns = sheet_config.get('columns', [])
        
        if data_start_cell and data_sql:
            # Replace temp table name
            data_sql_executed = data_sql.replace('temp_a1', self.temp_table_name)
            
            try:
                self.log_step("data_table", "start", f"Executing data query", sql=data_sql_executed)
                results = DatabaseManager.execute_sql_on_temp(data_sql_executed)
                
                if results:
                    self.log_step("data_table", "ok", f"Got {len(results)} rows for data table")
                    # Parse start cell
                    import re
                    match = re.match(r'([A-Z]+)(\d+)', data_start_cell)
                    if match:
                        start_col_letter = match.group(1)
                        start_row = int(match.group(2))
                        
                        for row_idx, row_data in enumerate(results):
                            values = list(row_data.values())
                            for col_idx, col_letter in enumerate(columns):
                                if col_idx < len(values):
                                    cell_ref = f"{col_letter}{start_row + row_idx}"
                                    ws[cell_ref] = values[col_idx]
                else:
                    self.log_step("data_table", "warning", "No data rows returned")
            except Exception as e:
                self.log_step("data_table", "error", f"Data table query failed: {str(e)}", sql=data_sql_executed)
        
        self.log_step("process_sheet", "ok", f"Completed sheet: {sheet_name}")
    
    def save_a1_csv(self, a1_df: pd.DataFrame) -> str:
        """
        Save A1 DataFrame to CSV
        
        Args:
            a1_df: A1 result DataFrame
        
        Returns:
            Path to saved CSV
        """
        output_path = self.export_path / "A1_result.csv"
        dataframe_to_csv(a1_df, output_path)
        return str(output_path)
    
    def save_a2_csv(self, a2_df: pd.DataFrame) -> Optional[str]:
        """
        Save A2 DataFrame to CSV
        
        Args:
            a2_df: A2 result DataFrame
        
        Returns:
            Path to saved CSV or None if empty
        """
        if a2_df.empty:
            return None
        
        output_path = self.export_path / "A2_result.csv"
        dataframe_to_csv(a2_df, output_path)
        return str(output_path)
