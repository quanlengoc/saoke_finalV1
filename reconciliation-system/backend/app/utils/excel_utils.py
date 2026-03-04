"""
Excel utility functions
Reading and writing Excel files with openpyxl
"""

import os
from pathlib import Path
from typing import Dict, Any, List, Optional, Union

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string

from app.core.config import get_storage_path


def column_letter_to_index(letter: str) -> int:
    """
    Convert Excel column letter to 0-based index
    
    Args:
        letter: Column letter (A, B, C, ..., AA, AB, ...)
    
    Returns:
        0-based column index
    """
    return column_index_from_string(letter.upper()) - 1


def index_to_column_letter(index: int) -> str:
    """
    Convert 0-based index to Excel column letter
    
    Args:
        index: 0-based column index
    
    Returns:
        Column letter
    """
    return get_column_letter(index + 1)


def read_excel_with_config(
    file_path: Union[str, Path],
    config: Dict[str, Any]
) -> pd.DataFrame:
    """
    Read Excel file with configuration
    
    Args:
        file_path: Path to Excel file
        config: File configuration dict with:
            - header_row: Row number containing headers (1-based)
            - data_start_row: Row number where data starts (1-based)
            - columns: Dict mapping internal names to Excel columns
            - sheet_name: Optional sheet name
    
    Returns:
        Pandas DataFrame with renamed columns
    """
    header_row = config.get('header_row', 1)
    data_start_row = config.get('data_start_row', 2)
    columns_map = config.get('columns', {})
    sheet_name = config.get('sheet_name', 0)  # Default to first sheet
    
    # Calculate skiprows (rows before data_start_row, excluding header)
    # pandas header is 0-based, so header_row-1
    # skiprows are rows to skip after header
    
    # Determine engine based on file extension
    file_ext = str(file_path).lower()
    engine = 'pyxlsb' if file_ext.endswith('.xlsb') else None
    
    # Read Excel file
    df = pd.read_excel(
        file_path,
        sheet_name=sheet_name,
        header=header_row - 1,  # 0-based
        skiprows=range(header_row, data_start_row - 1) if data_start_row > header_row + 1 else None,
        engine=engine
    )
    
    # If columns_map is provided, select and rename columns
    if columns_map:
        # Convert column letters to indices and select
        selected_cols = {}
        for internal_name, col_letter in columns_map.items():
            col_idx = column_letter_to_index(col_letter)
            if col_idx < len(df.columns):
                original_col = df.columns[col_idx]
                selected_cols[original_col] = internal_name
        
        # Select only mapped columns
        cols_to_select = list(selected_cols.keys())
        df = df[cols_to_select]
        
        # Rename columns
        df = df.rename(columns=selected_cols)
    
    return df


def read_csv_with_config(
    file_path: Union[str, Path],
    config: Dict[str, Any]
) -> pd.DataFrame:
    """
    Read CSV file with configuration
    
    Args:
        file_path: Path to CSV file
        config: File configuration dict
    
    Returns:
        Pandas DataFrame with renamed columns
    """
    header_row = config.get('header_row', 1)
    data_start_row = config.get('data_start_row', 2)
    columns_map = config.get('columns', {})
    
    # Read CSV file
    df = pd.read_csv(
        file_path,
        header=header_row - 1,
        skiprows=range(header_row, data_start_row - 1) if data_start_row > header_row + 1 else None
    )
    
    # If columns_map is provided, select and rename columns
    if columns_map:
        selected_cols = {}
        for internal_name, col_letter in columns_map.items():
            col_idx = column_letter_to_index(col_letter)
            if col_idx < len(df.columns):
                original_col = df.columns[col_idx]
                selected_cols[original_col] = internal_name
        
        cols_to_select = list(selected_cols.keys())
        df = df[cols_to_select]
        df = df.rename(columns=selected_cols)
    
    return df


def fill_excel_template(
    template_path: Union[str, Path],
    output_path: Union[str, Path],
    data: Dict[str, Any],
    sheet_configs: List[Dict[str, Any]]
) -> None:
    """
    Fill an Excel template with data
    
    Args:
        template_path: Path to template file
        output_path: Path for output file
        data: Data dictionary for static cell values
        sheet_configs: List of sheet configurations
    """
    # Load template
    wb = load_workbook(template_path)
    
    for sheet_config in sheet_configs:
        sheet_name = sheet_config.get('sheet_name')
        if sheet_name not in wb.sheetnames:
            continue
        
        ws = wb[sheet_name]
        
        # Fill static cells
        static_cells = sheet_config.get('static_cells', {})
        for cell_ref, value_template in static_cells.items():
            # Replace placeholders with actual values
            value = value_template
            for key, val in data.items():
                value = value.replace(f"{{{key}}}", str(val))
            ws[cell_ref] = value
        
        # Fill SQL-based cells (values should be pre-computed)
        sql_cells = sheet_config.get('sql_cells', [])
        for sql_cell in sql_cells:
            cell_ref = sql_cell.get('cell')
            value = sql_cell.get('value')  # Pre-computed value
            if cell_ref and value is not None:
                ws[cell_ref] = value
        
        # Fill data table
        data_start_cell = sheet_config.get('data_start_cell')
        data_rows = sheet_config.get('data_rows', [])
        columns = sheet_config.get('columns', [])
        
        if data_start_cell and data_rows:
            # Parse start cell
            import re
            match = re.match(r'([A-Z]+)(\d+)', data_start_cell)
            if match:
                start_col = match.group(1)
                start_row = int(match.group(2))
                
                for row_idx, row_data in enumerate(data_rows):
                    for col_idx, col_letter in enumerate(columns):
                        cell_ref = f"{col_letter}{start_row + row_idx}"
                        if isinstance(row_data, dict):
                            # row_data is a dict, get value by column index
                            keys = list(row_data.keys())
                            if col_idx < len(keys):
                                ws[cell_ref] = row_data[keys[col_idx]]
                        elif isinstance(row_data, (list, tuple)):
                            if col_idx < len(row_data):
                                ws[cell_ref] = row_data[col_idx]
    
    # Save output
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def get_template_path(template_name: str) -> Path:
    """
    Get full path to a template file
    
    Args:
        template_name: Template filename or relative path
    
    Returns:
        Absolute path to template
    """
    templates_dir = get_storage_path('templates')
    return templates_dir / template_name


def dataframe_to_excel(
    df: pd.DataFrame,
    output_path: Union[str, Path],
    sheet_name: str = "Sheet1"
) -> None:
    """
    Save DataFrame to Excel file
    
    Args:
        df: Pandas DataFrame
        output_path: Output file path
        sheet_name: Sheet name
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    df.to_excel(output_path, sheet_name=sheet_name, index=False)


def dataframe_to_csv(
    df: pd.DataFrame,
    output_path: Union[str, Path]
) -> None:
    """
    Save DataFrame to CSV file
    
    Args:
        df: Pandas DataFrame
        output_path: Output file path
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    df.to_csv(output_path, index=False, encoding='utf-8-sig')
